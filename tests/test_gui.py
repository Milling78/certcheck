"""Unit tests for the GUI's pure logic — config persistence, row mapping, export.
The Tkinter UI itself isn't exercised (no display); the testable helpers are."""
import os
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

pytest.importorskip("tkinter")           # skip cleanly on headless boxes without Tk
import certcheck_gui as g                  # noqa: E402
from certcheck import CertResult, Status   # noqa: E402


def _results():
    return [
        CertResult("a.com", 443, int(Status.OK), cn="a.com", issuer="R3",
                   not_after="2026-09-01T00:00:00Z", days_remaining=60, sans=["DNS:a.com"]),
        CertResult("b.com", 443, int(Status.CRITICAL), error="timed out"),
    ]


def test_results_to_rows_maps_fields():
    rows = g.results_to_rows(_results())
    assert rows[0]["status"] == "OK"
    assert rows[0]["host"] == "a.com"
    assert rows[0]["days_remaining"] == 60
    assert rows[0]["sans"] == "DNS:a.com"
    assert rows[1]["status"] == "CRITICAL"
    assert rows[1]["days_remaining"] == ""       # None -> blank for display/export
    assert rows[1]["error"] == "timed out"


def test_config_roundtrip(tmp_path, monkeypatch):
    monkeypatch.setenv("APPDATA", str(tmp_path))
    g.save_config({"hosts": "a.com\nb.com:8443", "warn": 45, "crit": 10, "timeout": 4.0})
    cfg = g.load_config()
    assert cfg["hosts"] == "a.com\nb.com:8443"
    assert cfg["warn"] == 45
    assert cfg["crit"] == 10
    assert cfg["timeout"] == 4.0


def test_load_config_missing_returns_empty(tmp_path, monkeypatch):
    monkeypatch.setenv("APPDATA", str(tmp_path / "does_not_exist_yet"))
    assert g.load_config() == {}


def test_export_xlsx_writes_workbook(tmp_path):
    pytest.importorskip("openpyxl")
    out = tmp_path / "out.xlsx"
    g.export_xlsx(_results(), str(out))
    assert out.exists() and out.stat().st_size > 0
    from openpyxl import load_workbook
    ws = load_workbook(str(out)).active
    assert ws["A1"].value == "Status"        # header row
    assert ws["B2"].value == "a.com"         # first data row, host column
    assert ws.max_row == 3                    # header + 2 results
