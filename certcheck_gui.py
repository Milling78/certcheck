#!/usr/bin/env python3
r"""certcheck GUI — point-and-click TLS certificate expiry checker.

A thin Tkinter front end over certcheck.py: it remembers your host list between
runs, scans on a button click, shows colour-coded results, and exports them to an
Excel (.xlsx) workbook. No command line required.

Run from source:   python certcheck_gui.py
Build an .exe:      build.ps1  ->  dist\certcheck-gui.exe
"""
from __future__ import annotations

import json
import os
import threading
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

from certcheck import CertResult, Status, __version__, check_many, parse_target

APP_NAME = "certcheck"
_DEFAULTS = {"hosts": "", "warn": 30, "crit": 7, "timeout": 5.0}


# ── config persistence (%APPDATA%\certcheck\config.json) ──────────────────────

def _config_dir() -> str:
    base = os.environ.get("APPDATA") or os.path.expanduser("~")
    d = os.path.join(base, APP_NAME)
    os.makedirs(d, exist_ok=True)
    return d


def _config_path() -> str:
    return os.path.join(_config_dir(), "config.json")


def load_config() -> dict:
    try:
        with open(_config_path(), encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def save_config(cfg: dict) -> None:
    try:
        with open(_config_path(), "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
    except OSError:
        pass


# ── pure helpers (testable) ───────────────────────────────────────────────────

_COLS = [
    ("status", "Status", 70),
    ("host", "Host", 190),
    ("port", "Port", 55),
    ("days_remaining", "Days", 55),
    ("not_after", "Expires (UTC)", 165),
    ("cn", "Common Name", 190),
    ("issuer", "Issuer", 170),
    ("error", "Error", 220),
]
_XLSX_EXTRA = ("sans", "SANs", 40)


def results_to_rows(results: list[CertResult]) -> list[dict]:
    """Flatten CertResult objects into display/export rows."""
    rows = []
    for r in results:
        rows.append({
            "status": Status(r.status).name,
            "host": r.host,
            "port": r.port,
            "days_remaining": "" if r.days_remaining is None else r.days_remaining,
            "not_after": r.not_after,
            "cn": r.cn,
            "issuer": r.issuer,
            "error": r.error,
            "sans": ", ".join(r.sans),
        })
    return rows


def export_xlsx(results: list[CertResult], path: str) -> None:
    """Write results to an .xlsx workbook with a header row and status colouring."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    fills = {
        "OK":       PatternFill("solid", fgColor="C6EFCE"),
        "WARNING":  PatternFill("solid", fgColor="FFEB9C"),
        "CRITICAL": PatternFill("solid", fgColor="FFC7CE"),
    }
    cols = _COLS + [_XLSX_EXTRA]

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificates"
    ws.append([c[1] for c in cols])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in results_to_rows(results):
        ws.append([row[c[0]] for c in cols])
        fill = fills.get(row["status"])
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    for i, (_, _, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(8, width // 6)
    ws.freeze_panes = "A2"
    wb.save(path)


# ── GUI ───────────────────────────────────────────────────────────────────────

_ROW_TAG = {"OK": ("ok", "#1a7f37"), "WARNING": ("warn", "#9a6700"), "CRITICAL": ("crit", "#b3261e")}


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.results: list[CertResult] = []
        self._scanning = False
        root.title(f"certcheck {__version__}")
        root.geometry("960x640")
        root.minsize(720, 480)

        cfg = {**_DEFAULTS, **load_config()}

        # ── top: hosts + settings ────────────────────────────────────────────
        top = ttk.Frame(root, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Hosts to check (one per line — host or host:port):").grid(
            row=0, column=0, columnspan=6, sticky="w")
        self.hosts_text = tk.Text(top, height=6, width=60, wrap="none", font=("Consolas", 10))
        self.hosts_text.grid(row=1, column=0, columnspan=6, sticky="ew", pady=(2, 8))
        self.hosts_text.insert("1.0", cfg.get("hosts", ""))
        top.columnconfigure(0, weight=1)

        ttk.Label(top, text="Warn ≤ days:").grid(row=2, column=0, sticky="e")
        self.warn = tk.IntVar(value=int(cfg.get("warn", 30)))
        ttk.Spinbox(top, from_=1, to=3650, textvariable=self.warn, width=6).grid(row=2, column=1, sticky="w", padx=(4, 14))

        ttk.Label(top, text="Critical ≤ days:").grid(row=2, column=2, sticky="e")
        self.crit = tk.IntVar(value=int(cfg.get("crit", 7)))
        ttk.Spinbox(top, from_=1, to=3650, textvariable=self.crit, width=6).grid(row=2, column=3, sticky="w", padx=(4, 14))

        ttk.Label(top, text="Timeout (s):").grid(row=2, column=4, sticky="e")
        self.timeout = tk.DoubleVar(value=float(cfg.get("timeout", 5.0)))
        ttk.Spinbox(top, from_=1, to=60, increment=1, textvariable=self.timeout, width=6).grid(row=2, column=5, sticky="w", padx=(4, 0))

        # ── buttons ───────────────────────────────────────────────────────────
        btns = ttk.Frame(root, padding=(10, 0))
        btns.pack(fill="x")
        self.scan_btn = ttk.Button(btns, text="Scan", command=self.scan)
        self.scan_btn.pack(side="left")
        self.export_btn = ttk.Button(btns, text="Export to Excel…", command=self.export, state="disabled")
        self.export_btn.pack(side="left", padx=6)

        # ── results table ─────────────────────────────────────────────────────
        mid = ttk.Frame(root, padding=10)
        mid.pack(fill="both", expand=True)
        cols = [c[0] for c in _COLS]
        self.tree = ttk.Treeview(mid, columns=cols, show="headings", selectmode="browse")
        for key, title, width in _COLS:
            self.tree.heading(key, text=title)
            self.tree.column(key, width=width, anchor="w", stretch=(key in ("host", "cn", "issuer", "error")))
        for tag, color in _ROW_TAG.values():
            self.tree.tag_configure(tag, foreground=color)
        ysb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        xsb = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        mid.rowconfigure(0, weight=1)
        mid.columnconfigure(0, weight=1)

        # ── status bar ────────────────────────────────────────────────────────
        self.status = tk.StringVar(value="Ready. Enter hosts and click Scan.")
        ttk.Label(root, textvariable=self.status, relief="sunken", anchor="w",
                  padding=(8, 3)).pack(fill="x", side="bottom")

        root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── actions ──────────────────────────────────────────────────────────────
    def _gather(self) -> dict:
        return {
            "hosts": self.hosts_text.get("1.0", "end").strip(),
            "warn": int(self.warn.get()),
            "crit": int(self.crit.get()),
            "timeout": float(self.timeout.get()),
        }

    def scan(self):
        if self._scanning:
            return
        cfg = self._gather()
        save_config(cfg)
        targets = [t for t in (parse_target(line) for line in cfg["hosts"].splitlines()) if t]
        if not targets:
            messagebox.showinfo("certcheck", "Enter at least one host to check.")
            return
        self._scanning = True
        self.scan_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.status.set(f"Scanning {len(targets)} host(s)…")
        self.tree.delete(*self.tree.get_children())

        def worker():
            try:
                results = check_many(targets, timeout=cfg["timeout"],
                                     warn_days=cfg["warn"], crit_days=cfg["crit"])
            except Exception as e:                              # noqa: BLE001
                self.root.after(0, lambda: self._scan_failed(str(e)))
                return
            self.root.after(0, lambda: self._show_results(results))

        threading.Thread(target=worker, daemon=True).start()

    def _scan_failed(self, msg: str):
        self._scanning = False
        self.scan_btn.config(state="normal")
        self.status.set("Scan failed.")
        messagebox.showerror("certcheck", f"Scan failed:\n{msg}")

    def _show_results(self, results: list[CertResult]):
        self.results = results
        for row in results_to_rows(results):
            tag = _ROW_TAG[row["status"]][0]
            self.tree.insert("", "end", tags=(tag,), values=[row[c[0]] for c in _COLS])
        ok = sum(1 for r in results if r.status == Status.OK)
        warn = sum(1 for r in results if r.status == Status.WARNING)
        crit = sum(1 for r in results if r.status == Status.CRITICAL)
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        self.status.set(f"{len(results)} checked — {ok} OK, {warn} warning, {crit} critical   (scanned {stamp})")
        self._scanning = False
        self.scan_btn.config(state="normal")
        self.export_btn.config(state=("normal" if results else "disabled"))

    def export(self):
        if not self.results:
            return
        default = f"certcheck_{datetime.now():%Y%m%d_%H%M}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Export to Excel", defaultextension=".xlsx",
            initialfile=default, filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return
        try:
            export_xlsx(self.results, path)
        except ImportError:
            messagebox.showerror("certcheck", "Excel export needs the 'openpyxl' package.\n\npip install openpyxl")
            return
        except Exception as e:                                  # noqa: BLE001
            messagebox.showerror("certcheck", f"Could not write the file:\n{e}")
            return
        self.status.set(f"Exported {len(self.results)} row(s) to {os.path.basename(path)}")

    def _on_close(self):
        save_config(self._gather())
        self.root.destroy()


def main() -> int:
    root = tk.Tk()
    try:
        ttk.Style().theme_use("vista")   # native-ish on Windows; harmless elsewhere
    except tk.TclError:
        pass
    App(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
